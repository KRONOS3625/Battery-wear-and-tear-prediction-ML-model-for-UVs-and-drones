from __future__ import annotations

import io
import json
from http import HTTPStatus
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import urlparse

import joblib
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


ROOT = Path(__file__).resolve().parent
WEB_DIR = ROOT / "web"
MODEL_BUNDLE = ROOT / "models" / "battery_health_models.joblib"
MODEL_METADATA = ROOT / "models" / "model_metadata.json"

bundle = joblib.load(MODEL_BUNDLE)
with MODEL_METADATA.open("r", encoding="utf-8") as handle:
    metadata = json.load(handle)

FEATURE_NAMES = bundle["feature_names"]
soh_model = bundle["soh_model"]
rul_model = bundle["rul_model"]
anomaly_model = bundle["anomaly_model"]

VEHICLE_FACTORS = {
    "ev": {"wear": 1.0, "temp_bias": 0.0, "load": 0.14},
    "drone": {"wear": 1.12, "temp_bias": 2.0, "load": 0.24},
}
CHEMISTRY_FACTORS = {
    "li-ion": {"wear": 1.0, "safe_soc": 0.82},
    "lipo": {"wear": 1.08, "safe_soc": 0.78},
    "lfp": {"wear": 0.92, "safe_soc": 0.9},
}


def clamp(value: float, low: float, high: float) -> float:
    return max(low, min(high, value))


def scale_risk(value: float, minimum: float, maximum: float, invert: bool = False) -> float:
    if maximum <= minimum:
        return 0.0
    score = (value - minimum) / (maximum - minimum)
    score = clamp(score, 0.0, 1.0)
    return (1.0 - score) if invert else score


def sigmoid(value: float) -> float:
    return 1.0 / (1.0 + np.exp(-value))


def tree_spread(model: object, feature_vector: np.ndarray) -> float:
    estimators = getattr(model, "estimators_", None)
    if estimators is None:
        return 0.0
    predictions = np.asarray([est.predict(feature_vector)[0] for est in estimators], dtype=float)
    return float(np.std(predictions))


def health_band(soh: float) -> tuple[str, str]:
    if soh >= 92:
        return "Stable", "The pack is operating with healthy remaining margin."
    if soh >= 84:
        return "Monitor", "Wear is visible, but the battery is still serviceable."
    if soh >= 75:
        return "Elevated", "Degradation is rising and maintenance planning should begin."
    return "Critical", "The battery is nearing end-of-life and needs intervention."


def thermal_label(temp_c: float) -> tuple[str, str]:
    if temp_c <= 32:
        return "Low", "Thermal stress is in a comfortable operating band."
    if temp_c <= 40:
        return "Moderate", "Heat is contributing to aging and should be controlled."
    return "High", "Thermal stress is strongly accelerating degradation."


def derive_model_features(payload: dict[str, object]) -> tuple[np.ndarray, dict[str, float]]:
    capacity_ah = float(payload["capacity_ah"])
    current_a = float(payload["current_a"])
    dod_percent = float(payload["dod_percent"])
    voltage_v = float(payload["voltage_v"])
    internal_resistance_ohm = float(payload["internal_resistance_mohm"]) / 1000.0
    cycle_number = float(payload["cycle_number"])
    temperature_c = float(payload["temperature_c"])
    ambient_temperature_c = float(payload["ambient_temperature_c"])

    avg_voltage_v = voltage_v
    min_voltage_v = max(2.2, voltage_v - (0.06 + current_a * 0.045 + internal_resistance_ohm * 12.0))
    avg_current_a = current_a
    max_current_a = current_a * 1.18
    current_std_a = max(0.04, current_a * 0.12)
    discharge_time_s = max(180.0, (capacity_ah * (dod_percent / 100.0) / max(current_a, 0.2)) * 3600.0)
    energy_wh = avg_voltage_v * capacity_ah * (dod_percent / 100.0)
    load_c_rate = avg_current_a / max(capacity_ah, 0.1)

    feature_map = {
        "internal_resistance_ohm": internal_resistance_ohm,
        "capacity_ah": capacity_ah,
        "cycle_number": cycle_number,
        "temperature_c": temperature_c,
        "ambient_temperature_c": ambient_temperature_c,
        "avg_voltage_v": avg_voltage_v,
        "min_voltage_v": min_voltage_v,
        "avg_current_a": avg_current_a,
        "max_current_a": max_current_a,
        "current_std_a": current_std_a,
        "discharge_time_s": discharge_time_s,
        "energy_wh": energy_wh,
        "load_c_rate": load_c_rate,
    }
    vector = np.asarray([[feature_map[name] for name in FEATURE_NAMES]], dtype=float)
    return vector, feature_map


def contextual_stress(payload: dict[str, object], feature_map: dict[str, float]) -> tuple[dict[str, float], float]:
    vehicle = str(payload["vehicle_type"])
    chemistry = str(payload["chemistry"])
    vehicle_factor = VEHICLE_FACTORS[vehicle]
    chemistry_factor = CHEMISTRY_FACTORS[chemistry]

    soc = float(payload["soc_percent"])
    dod = float(payload["dod_percent"])
    rest_hours = float(payload["rest_hours"])
    age_days = float(payload["age_days"])
    fast_charge_count = float(payload["fast_charge_count"])
    payload_mass = float(payload["payload_kg"])
    trip_distance = float(payload["trip_distance_km"])
    average_speed = float(payload["average_speed_kmh"])
    vibration_g = float(payload["vibration_g"])
    ambient_temperature_c = float(payload["ambient_temperature_c"])
    pack_series_cells = float(payload["pack_series_cells"])

    soc_risk = clamp(abs(soc / 100.0 - chemistry_factor["safe_soc"]) / 0.45, 0.0, 1.0)
    dod_risk = clamp((dod - 55.0) / 45.0, 0.0, 1.0)
    rest_risk = clamp((1.0 - min(rest_hours, 6.0) / 6.0), 0.0, 1.0)
    calendar_risk = clamp(age_days / 2200.0, 0.0, 1.0)
    fast_charge_risk = clamp(fast_charge_count / 260.0, 0.0, 1.0)
    ambient_risk = clamp(abs(ambient_temperature_c - 25.0) / 25.0, 0.0, 1.0)
    voltage_sag_risk = clamp((feature_map["avg_voltage_v"] - feature_map["min_voltage_v"]) / 0.7, 0.0, 1.0)
    c_rate_risk = clamp((feature_map["load_c_rate"] - 0.9) / 2.0, 0.0, 1.0)
    payload_risk = clamp(payload_mass / 12.0, 0.0, 1.0)
    distance_risk = clamp(trip_distance / 160.0, 0.0, 1.0)
    speed_risk = clamp(average_speed / 120.0, 0.0, 1.0)
    vibration_risk = clamp(vibration_g / 4.0, 0.0, 1.0)
    pack_imbalance_risk = clamp((pack_series_cells - 12.0) / 10.0, 0.0, 1.0)

    contributions = {
        "Internal resistance": 0.14 + scale_risk(
            feature_map["internal_resistance_ohm"],
            metadata["feature_ranges"]["internal_resistance_ohm"]["p10"],
            metadata["feature_ranges"]["internal_resistance_ohm"]["p90"],
        ),
        "Capacity fade": 0.14 + scale_risk(
            feature_map["capacity_ah"],
            metadata["feature_ranges"]["capacity_ah"]["p10"],
            metadata["feature_ranges"]["capacity_ah"]["p90"],
            invert=True,
        ),
        "Cycle aging": 0.14 + scale_risk(
            feature_map["cycle_number"],
            metadata["feature_ranges"]["cycle_number"]["p10"],
            metadata["feature_ranges"]["cycle_number"]["p90"],
        ),
        "Thermal load": 0.12 + max(
            scale_risk(feature_map["temperature_c"], 28.0, 48.0),
            ambient_risk,
        ),
        "State of charge": 0.08 + soc_risk,
        "Depth of discharge": 0.08 + dod_risk,
        "Charge severity": 0.08 + max(c_rate_risk, fast_charge_risk),
        "Mission stress": 0.08 + max(
            vehicle_factor["load"] * payload_risk + 0.4 * vibration_risk,
            0.45 * distance_risk + 0.35 * speed_risk,
            pack_imbalance_risk,
            voltage_sag_risk,
        ),
        "Calendar aging": 0.08 + max(calendar_risk, rest_risk * 0.5),
    }

    stress_index = (
        0.14 * soc_risk
        + 0.12 * dod_risk
        + 0.10 * rest_risk
        + 0.12 * calendar_risk
        + 0.12 * fast_charge_risk
        + 0.10 * ambient_risk
        + 0.08 * voltage_sag_risk
        + 0.08 * c_rate_risk
        + 0.08 * payload_risk
        + 0.06 * distance_risk
        + 0.05 * speed_risk
        + 0.05 * vibration_risk
    )
    stress_index *= vehicle_factor["wear"] * chemistry_factor["wear"]
    return contributions, clamp(stress_index, 0.0, 1.35)


def build_forecast(
    cycle_number: float,
    capacity_ah: float,
    rul_cycles: float,
    initial_capacity_ah: float,
    stress_index: float,
) -> dict[str, list[float] | float]:
    eol_capacity = initial_capacity_ah * 0.70
    horizon = max(int(round(rul_cycles)), 1)
    future_steps = np.linspace(0.0, 1.0, 28)
    future_cycles = cycle_number + future_steps * horizon

    if capacity_ah <= eol_capacity:
        future_capacity = np.full_like(future_cycles, capacity_ah)
    else:
        curve_power = 1.05 + stress_index * 0.7
        future_capacity = capacity_ah - (capacity_ah - eol_capacity) * np.power(future_steps, curve_power)

    history_length = min(max(int(cycle_number // 16), 6), 10)
    history_steps = np.linspace(0.0, 1.0, history_length)
    history_cycles = np.maximum(cycle_number - history_steps[::-1] * max(cycle_number * 0.5, 1.0), 0.0)
    history_capacity = initial_capacity_ah - (initial_capacity_ah - capacity_ah) * np.power(history_steps, 1.12)

    return {
        "history_cycles": [float(round(value, 2)) for value in history_cycles],
        "history_capacity": [float(round(value, 4)) for value in history_capacity],
        "future_cycles": [float(round(value, 2)) for value in future_cycles],
        "future_capacity": [float(round(value, 4)) for value in future_capacity],
        "eol_capacity": float(round(eol_capacity, 4)),
    }


def build_pack_heatmap(base_soh: float, stress_index: float, pack_series_cells: int) -> list[dict[str, float | int]]:
    rng = np.random.default_rng(42 + pack_series_cells)
    spread = 0.7 + stress_index * 5.0
    cells = []
    for index in range(pack_series_cells):
        offset = rng.normal(0.0, spread) + (0.6 if index % 5 == 0 else 0.0) * stress_index
        cell_soh = clamp(base_soh - offset, 55.0, 100.0)
        cells.append({"cell": index + 1, "soh": round(cell_soh, 2)})
    return cells


def build_uncertainty(soh_spread: float, rul_spread: float, anomaly_score: float) -> dict[str, float]:
    confidence = clamp(91.0 - soh_spread * 1.8 - rul_spread * 0.85 - max(0.0, -anomaly_score) * 18.0, 42.0, 97.0)
    score = round(1.0 - confidence / 100.0, 3)
    confidence = round(confidence, 1)
    return {"confidence_percent": confidence, "uncertainty_score": round(score, 3)}


def build_what_if(payload: dict[str, object], current_soh: float, current_rul: float) -> dict[str, float | str]:
    optimized = dict(payload)
    optimized["temperature_c"] = max(24.0, float(payload["temperature_c"]) - 6.0)
    optimized["ambient_temperature_c"] = max(22.0, float(payload["ambient_temperature_c"]) - 4.0)
    optimized["soc_percent"] = min(float(payload["soc_percent"]), 78.0)
    optimized["dod_percent"] = min(float(payload["dod_percent"]), 62.0)
    optimized["current_a"] = max(0.7, float(payload["current_a"]) * 0.85)
    optimized["fast_charge_count"] = max(0.0, float(payload["fast_charge_count"]) * 0.55)
    optimized["rest_hours"] = max(2.0, float(payload["rest_hours"]))
    optimized["payload_kg"] = max(0.0, float(payload["payload_kg"]) * 0.8)
    optimized["vibration_g"] = max(0.2, float(payload["vibration_g"]) * 0.7)
    optimized["voltage_v"] = min(float(payload["voltage_v"]) + 0.07, 4.2)

    optimized_vector, optimized_map = derive_model_features(optimized)
    optimized_soh = float(soh_model.predict(optimized_vector)[0])
    optimized_rul = float(rul_model.predict(optimized_vector)[0])
    _, optimized_stress = contextual_stress(optimized, optimized_map)
    optimized_soh = clamp(optimized_soh - optimized_stress * 12.0, 50.0, 100.0)
    optimized_rul = max(0.0, optimized_rul * (1.0 - optimized_stress * 0.42))
    optimized_soh = max(optimized_soh, min(100.0, current_soh + 2.4))
    optimized_rul = max(optimized_rul, current_rul * 1.12)

    return {
        "title": "Low-stress operating scenario",
        "vehicle_type": str(payload["vehicle_type"]).upper(),
        "projected_soh_percent": round(optimized_soh, 2),
        "projected_rul_cycles": round(optimized_rul, 1),
        "savings_summary": "Lower temperature, lower DoD, and fewer fast-charge events improve lifespan the most.",
    }


def maintenance_recommendations(
    payload: dict[str, object],
    soh: float,
    failure_probability: float,
    stress_index: float,
) -> list[str]:
    notes: list[str] = []
    if float(payload["temperature_c"]) > 38 or float(payload["ambient_temperature_c"]) > 36:
        notes.append("Reduce thermal loading with better cooling or shorter high-power runs.")
    if float(payload["soc_percent"]) > 88:
        notes.append("Avoid storing the pack at very high state of charge for long periods.")
    if float(payload["dod_percent"]) > 80:
        notes.append("Use shallower discharge windows to slow capacity fade.")
    if float(payload["fast_charge_count"]) > 80:
        notes.append("Cut down fast-charge frequency or lower charge current where possible.")
    if str(payload["vehicle_type"]) == "drone" and float(payload["payload_kg"]) > 4.0:
        notes.append("Reduce payload or shorten sorties to lower mission stress on the cells.")
    if str(payload["vehicle_type"]) == "ev" and float(payload["average_speed_kmh"]) > 90:
        notes.append("Sustained high-speed operation is increasing current draw and thermal load.")
    if stress_index > 0.55:
        notes.append("Schedule a balance check and inspect for cell-to-cell voltage imbalance.")
    if soh < 78 or failure_probability > 0.55:
        notes.append("Prioritize pack diagnostics and plan replacement or reconditioning.")
    if len(notes) < 3:
        notes.append("Track weekly SoH and temperature trends to catch degradation before it accelerates.")
    if len(notes) < 3:
        notes.append("Keep pack balancing and connector inspection in the maintenance schedule.")
    return notes[:5]


def build_report_workbook(payload: dict[str, object], prediction: dict[str, object]) -> bytes:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    accent_fill = PatternFill(fill_type="solid", fgColor="DCEBFA")
    header_font = Font(bold=True)

    def write_section(sheet, title: str, rows: list[tuple[str, object]], start_row: int) -> int:
        sheet.cell(row=start_row, column=1, value=title)
        sheet.cell(row=start_row, column=1).font = header_font
        sheet.cell(row=start_row, column=1).fill = accent_fill
        current = start_row + 1
        for key, value in rows:
            sheet.cell(row=current, column=1, value=key)
            sheet.cell(row=current, column=2, value=value)
            current += 1
        return current + 1

    outputs = prediction["outputs"]
    current_row = 1
    current_row = write_section(
        summary,
        "Prediction Summary",
        [
            ("State of Health (%)", outputs["soh_percent"]),
            ("Wear (%)", outputs["wear_percent"]),
            ("Remaining Useful Life (cycles)", outputs["rul_cycles"]),
            ("Estimated End-of-Life Cycle", outputs["estimated_eol_cycle"]),
            ("Failure Probability (%)", outputs["failure_probability"]),
            ("Thermal Risk", outputs["thermal_label"]),
            ("Confidence (%)", outputs["confidence_percent"]),
            ("Stress Index", outputs["stress_index"]),
            ("Anomaly Flag", outputs["anomaly_flag"]),
        ],
        current_row,
    )
    current_row = write_section(
        summary,
        "What-If Scenario",
        [
            ("Scenario Title", prediction["what_if"]["title"]),
            ("Projected SoH (%)", prediction["what_if"]["projected_soh_percent"]),
            ("Projected RUL (cycles)", prediction["what_if"]["projected_rul_cycles"]),
            ("Notes", prediction["what_if"]["savings_summary"]),
        ],
        current_row,
    )

    inputs_sheet = workbook.create_sheet("Inputs")
    inputs_sheet.append(["Input", "Value"])
    inputs_sheet["A1"].font = header_font
    inputs_sheet["B1"].font = header_font
    for key, value in payload.items():
        inputs_sheet.append([key, value])

    features_sheet = workbook.create_sheet("Derived Features")
    features_sheet.append(["Feature", "Value"])
    features_sheet["A1"].font = header_font
    features_sheet["B1"].font = header_font
    for key, value in prediction["report"]["feature_map"].items():
        features_sheet.append([key, value])

    rec_sheet = workbook.create_sheet("Recommendations")
    rec_sheet.append(["Recommendation"])
    rec_sheet["A1"].font = header_font
    for note in prediction["recommendations"]:
        rec_sheet.append([note])

    heatmap_sheet = workbook.create_sheet("Pack Heatmap")
    heatmap_sheet.append(["Cell", "Predicted SoH (%)"])
    heatmap_sheet["A1"].font = header_font
    heatmap_sheet["B1"].font = header_font
    for cell in prediction["charts"]["pack_heatmap"]:
        heatmap_sheet.append([cell["cell"], cell["soh"]])

    contribution_sheet = workbook.create_sheet("Wear Drivers")
    contribution_sheet.append(["Driver", "Contribution (%)"])
    contribution_sheet["A1"].font = header_font
    contribution_sheet["B1"].font = header_font
    for item in prediction["charts"]["contributions"]:
        contribution_sheet.append([item["label"], item["value"]])

    for sheet in workbook.worksheets:
        sheet.column_dimensions["A"].width = 28
        sheet.column_dimensions["B"].width = 24
        sheet.column_dimensions["C"].width = 20

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def predict(payload: dict[str, object]) -> dict[str, object]:
    feature_vector, feature_map = derive_model_features(payload)
    base_soh = float(soh_model.predict(feature_vector)[0])
    base_rul = float(rul_model.predict(feature_vector)[0])
    anomaly_score = float(anomaly_model.decision_function(feature_vector)[0])
    _, anomaly_flag = anomaly_model.score_samples(feature_vector), anomaly_model.predict(feature_vector)[0]
    contributions, stress_index = contextual_stress(payload, feature_map)

    adjusted_soh = clamp(base_soh - stress_index * 12.0, 50.0, 100.0)
    adjusted_rul = max(0.0, base_rul * (1.0 - stress_index * 0.42))

    soh_spread = tree_spread(soh_model, feature_vector)
    rul_spread = tree_spread(rul_model, feature_vector)
    uncertainty = build_uncertainty(soh_spread, rul_spread, anomaly_score)

    wear_percent = clamp(100.0 - adjusted_soh, 0.0, 100.0)
    failure_probability = clamp(
        float(
            sigmoid(
                -1.9
                + wear_percent / 12.0
                + stress_index * 2.4
                + max(0.0, -anomaly_score) * 2.2
                + (float(payload["temperature_c"]) - 32.0) / 10.0
            )
        ),
        0.01,
        0.99,
    )
    state_label, state_text = health_band(adjusted_soh)
    thermal_state, thermal_text = thermal_label(float(payload["temperature_c"]))
    baseline_capacity = float(metadata["baseline"]["mean_initial_capacity_ah"])
    pack_cells = int(float(payload["pack_series_cells"]))
    heatmap = build_pack_heatmap(adjusted_soh, stress_index, pack_cells)
    forecast = build_forecast(float(payload["cycle_number"]), float(payload["capacity_ah"]), adjusted_rul, baseline_capacity, stress_index)
    what_if = build_what_if(payload, adjusted_soh, adjusted_rul)
    recommendations = maintenance_recommendations(payload, adjusted_soh, failure_probability, stress_index)

    contribution_colors = ["#f97316", "#0f766e", "#2563eb", "#ef4444", "#f59e0b", "#14b8a6", "#8b5cf6", "#22c55e", "#fb7185"]
    weighted_sum = sum(contributions.values()) or 1.0
    contribution_payload = []
    for index, (label, value) in enumerate(contributions.items()):
        contribution_payload.append(
            {
                "label": label,
                "value": round(value / weighted_sum * 100.0, 1),
                "color": contribution_colors[index % len(contribution_colors)],
            }
        )

    return {
        "outputs": {
            "soh_percent": round(adjusted_soh, 2),
            "wear_percent": round(wear_percent, 2),
            "rul_cycles": round(adjusted_rul, 1),
            "estimated_eol_cycle": round(float(payload["cycle_number"]) + adjusted_rul, 1),
            "degradation_rate_per_cycle": round(wear_percent / max(float(payload["cycle_number"]), 1.0), 4),
            "failure_probability": round(failure_probability * 100.0, 1),
            "health_label": state_label,
            "health_text": state_text,
            "thermal_label": thermal_state,
            "thermal_text": thermal_text,
            "confidence_percent": uncertainty["confidence_percent"],
            "anomaly_flag": bool(anomaly_flag == -1),
            "stress_index": round(stress_index * 100.0, 1),
        },
        "charts": {
            "forecast": forecast,
            "contributions": contribution_payload,
            "pack_heatmap": heatmap,
        },
        "what_if": what_if,
        "recommendations": recommendations,
        "report": {
            "vehicle_type": payload["vehicle_type"],
            "chemistry": payload["chemistry"],
            "base_model_soh": round(base_soh, 2),
            "base_model_rul": round(base_rul, 1),
            "feature_map": {key: round(value, 4) for key, value in feature_map.items()},
        },
        "meta": {
            "model_metrics": metadata["metrics"],
            "model_comparison": metadata["model_comparison"],
            "battery_count": metadata["battery_count"],
            "sample_count": metadata["sample_count"],
        },
    }


class BatteryHandler(SimpleHTTPRequestHandler):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=str(WEB_DIR), **kwargs)

    def _send_json(self, payload: dict[str, object], status: int = HTTPStatus.OK) -> None:
        body = json.dumps(payload).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def do_GET(self) -> None:
        parsed = urlparse(self.path)
        if parsed.path == "/api/metadata":
            self._send_json(
                {
                    "sample_count": metadata["sample_count"],
                    "battery_count": metadata["battery_count"],
                    "metrics": metadata["metrics"],
                    "sources": metadata["sources"],
                    "ui_bounds": metadata["ui_bounds"],
                    "model_comparison": metadata["model_comparison"],
                }
            )
            return

        if parsed.path == "/api/report.xlsx":
            self._send_json({"error": "Use POST /api/report.xlsx with the current inputs."}, status=HTTPStatus.METHOD_NOT_ALLOWED)
            return

        if parsed.path in {"/", ""}:
            self.path = "/index.html"
        super().do_GET()

    def do_POST(self) -> None:
        endpoint = urlparse(self.path).path
        if endpoint not in {"/api/predict", "/api/report.xlsx"}:
            self._send_json({"error": "Unknown endpoint."}, status=HTTPStatus.NOT_FOUND)
            return
        try:
            content_length = int(self.headers.get("Content-Length", "0"))
            raw_body = self.rfile.read(content_length)
            payload = json.loads(raw_body.decode("utf-8"))
            response = predict(payload)
        except Exception as exc:  # noqa: BLE001
            self._send_json({"error": str(exc)}, status=HTTPStatus.BAD_REQUEST)
            return
        if endpoint == "/api/predict":
            self._send_json(response)
            return

        workbook_bytes = build_report_workbook(payload, response)
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.send_header("Content-Disposition", 'attachment; filename="battery_intelligence_report.xlsx"')
        self.send_header("Content-Length", str(len(workbook_bytes)))
        self.end_headers()
        self.wfile.write(workbook_bytes)


def main() -> None:
    host = "127.0.0.1"
    port = 8000
    server = ThreadingHTTPServer((host, port), BatteryHandler)
    print(f"Battery predictor running at http://{host}:{port}")
    server.serve_forever()


if __name__ == "__main__":
    main()
